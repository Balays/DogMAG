#!/usr/bin/env python3
"""Build a journal-style Supplementary Material workbook from DogMAG TSV tables.

The script reads every TSV file in supplementary_tables/ whose filename starts
with supplementary_table_<number>_ and writes one Excel worksheet per table.
Worksheet names are "Supplementary Table S1", "Supplementary Table S2", ... .
The TSV values are written as text so accessions, identifiers, dates, and
formatted numeric strings are preserved exactly as distributed.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TABLE_RE = re.compile(r"^supplementary_table_(\d+)_.*\.tsv$")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)


def supplementary_tables(input_dir: Path) -> list[tuple[int, Path]]:
    tables: list[tuple[int, Path]] = []
    for path in input_dir.iterdir():
        match = TABLE_RE.match(path.name)
        if match:
            tables.append((int(match.group(1)), path))
    return sorted(tables)


def iter_tsv_rows(path: Path) -> Iterable[list[str | None]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for row in reader:
            yield [cell if cell != "" else None for cell in row]


def safe_sheet_name(number: int) -> str:
    # 22 characters for S1-S9, comfortably under Excel's 31-character limit.
    return f"Supplementary Table S{number}"


def write_table_sheet(wb: Workbook, number: int, path: Path) -> None:
    ws = wb.create_sheet(title=safe_sheet_name(number))
    max_widths: list[int] = []
    row_count = 0
    col_count = 0

    for row in iter_tsv_rows(path):
        row_count += 1
        col_count = max(col_count, len(row))
        ws.append(row)
        # Estimate widths from the header and the first 200 data rows only.
        if row_count <= 201:
            for idx, value in enumerate(row):
                width = len(str(value)) if value is not None else 0
                if idx >= len(max_widths):
                    max_widths.append(width)
                else:
                    max_widths[idx] = max(max_widths[idx], width)

    if row_count == 0:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header = ws[1]
    for cell in header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.number_format = "@"

    for idx in range(1, col_count + 1):
        estimated = max_widths[idx - 1] if idx <= len(max_widths) else 10
        # Keep the workbook readable without creating extremely wide columns.
        ws.column_dimensions[get_column_letter(idx)].width = min(max(estimated + 2, 10), 40)


def update_checksum_file(checksum_path: Path, workbook_path: Path) -> None:
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    workbook_rel = workbook_path.as_posix()
    if checksum_path.exists():
        lines = checksum_path.read_text(encoding="utf-8").splitlines()
        lines = [line for line in lines if not line.endswith(f"\t{workbook_rel}")]
    else:
        lines = []
    lines.append(f"{digest}\t{workbook_rel}")
    checksum_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default="supplementary_tables")
    parser.add_argument("--output", default="supplementary_material.xlsx")
    parser.add_argument("--checksum-file", default="checksums/SHA256SUMS.tsv")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    checksum_path = Path(args.checksum_file)

    tables = supplementary_tables(input_dir)
    if not tables:
        raise SystemExit(f"No supplementary_table_<number>_*.tsv files found in {input_dir}")

    wb = Workbook()
    # Remove default sheet so the workbook contains only supplementary table tabs.
    default_sheet = wb.active
    wb.remove(default_sheet)
    wb.properties.title = "DogMAG Supplementary Material"
    wb.properties.subject = "DogMAG supplementary tables S1-S9"
    wb.properties.creator = "DogMAG workflow"

    for number, path in tables:
        write_table_sheet(wb, number, path)

    wb.save(output_path)
    update_checksum_file(checksum_path, output_path)
    print(f"Wrote {output_path} with {len(tables)} sheets")


if __name__ == "__main__":
    main()
