from __future__ import annotations

import argparse
import csv
import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook


STUDY = "PRJEB115259"
HEADERS = [
    "sample",
    "study",
    "instrument_model",
    "library_name",
    "library_source",
    "library_selection",
    "library_strategy",
    "library_layout",
    "file_name",
    "file_md5",
    "library_construction_protocol",
]

DMD_MD5 = {
    "DMD_Boszi_MN_bc08_pass.fastq.gz": "0c9d5aa8d363c98db578473581ce382b",
    "DMD_Boszi_Zymo_HMW_bc04_pass.fastq.gz": "90f0869bcc4d55d36a6fd888204fe867",
    "DMD_Brios_MN_bc07_pass.fastq.gz": "5963384186ffc823538cf9d01bf8bf65",
    "DMD_Brios_Zymo_HMW_bc03_pass.fastq.gz": "2355242f861cac982d096d42f9b19d03",
    "DMD_Loki_MN_bc05_pass.fastq.gz": "83beab89e6609da2c88b76e611afdffd",
    "DMD_Loki_Zymo_HMW_bc01_pass.fastq.gz": "24a52db438f92f52c6afdb7fa286f5c9",
    "DMD_Sugo_MN_bc06_pass.fastq.gz": "42445547c1719fbb9ee89001feece462",
    "DMD_Sugo_Zymo_HMW_bc02_pass.fastq.gz": "3bbcdd82c96d26f7f6f2a72b7da400cb",
}

SERTE_MD5 = {
    "SertePerti_ONT_WGS_concatenated_barcode01.fastq.gz": "cdb4d5002671cf19738852f1238efed6",
    "SertePerti_ONT_WGS_concatenated_barcode02.fastq.gz": "8cdf564d627c455f27990f24cc008f9a",
    "SertePerti_ONT_WGS_concatenated_barcode03.fastq.gz": "58bad1b0910c309ca8fb439234e6fe4e",
    "SertePerti_ONT_WGS_concatenated_barcode04.fastq.gz": "3faddf73d6b18084cf3e3e4f002cf53b",
    "SertePerti_ONT_WGS_concatenated_barcode05.fastq.gz": "f8f93d908dae9d8b3cb112c9628c8da5",
    "SertePerti_ONT_WGS_concatenated_barcode06.fastq.gz": "4e6ccce9557544a65fefb49b3f280ac8",
}

SERTE_BARCODE = {
    1: ("Rozi", "HMW"),
    2: ("Degesz", "HMW"),
    3: ("Ilka", "HMW"),
    4: ("Rozi", "ZymoBIOMICS_96_MagBead"),
    5: ("Degesz", "ZymoBIOMICS_96_MagBead"),
    6: ("Ilka", "ZymoBIOMICS_96_MagBead"),
}

EXCLUDED_ALL_KENNEL_BARCODES = {67, 68}


def safe_alias(value: str) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^A-Za-z0-9]+", "_", value)).strip("_")


def md5sum(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def md5sum_if_available(path: Path) -> str:
    try:
        if path.stat().st_size <= 0:
            return ""
        return md5sum(path)
    except OSError:
        return ""


def all_kennel_sample_map(workbook_path: Path) -> dict[int, str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["FASTQ mapping"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    index = {name: pos for pos, name in enumerate(headers)}
    result: dict[int, str] = {}
    for row in rows:
        if row[index["sequencing_type"]] != "long_read":
            continue
        if row[index["project"]] != "All_Kennels_WGS":
            continue
        barcode = row[index["barcode"]]
        sample_id = row[index["sample_id"]]
        if barcode and sample_id:
            result.setdefault(int(barcode), str(sample_id))
    workbook.close()
    return result


def all_kennel_asad_id_map(workbook_path: Path) -> dict[int, str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["WGS-Dogs information "]
    result: dict[int, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        barcode = row[11]
        sample_id = row[18]
        if barcode and 1 <= int(barcode) <= 72 and sample_id:
            result[int(barcode)] = str(sample_id).strip()
    workbook.close()

    # The workbook has one blank ID and one date typo, both resolvable from
    # collection date, kennel/dog code, and the recorded raw-data filename.
    result[57] = "SRTP_G_PF1_230226"
    result[65] = "PP_0_AM2_230714"
    return result


def make_row(
    *, sample: str, instrument: str, library_name: str, file_name: str,
    file_md5: str, protocol: str,
) -> dict[str, str]:
    return {
        "sample": sample,
        "study": STUDY,
        "instrument_model": instrument,
        "library_name": library_name,
        "library_source": "METAGENOMIC",
        "library_selection": "RANDOM",
        "library_strategy": "WGS",
        "library_layout": "SINGLE",
        "file_name": file_name,
        "file_md5": file_md5,
        "library_construction_protocol": protocol,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--canmag", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--all-kennel-dir", type=Path)
    parser.add_argument("--md5-from-manifest", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--require-complete", action="store_true")
    args = parser.parse_args()

    canmag = args.canmag.resolve()
    template = canmag / "ENA" / STUDY / "fastq1_template_1782115386885.tsv"
    mapping = canmag / "metadata" / "canmag_fastq_dog_mapping_FINAL.xlsx"
    all_kennel_dir = args.all_kennel_dir or canmag / "fastq" / "All_Kennel_ONT"
    output = args.output or template.with_name(f"fastq1_{STUDY}_long_reads_draft.tsv")

    cached_md5: dict[str, str] = {}
    if args.md5_from_manifest:
        with args.md5_from_manifest.open("r", encoding="utf-8-sig", newline="") as handle:
            handle.readline()
            cached_md5 = {
                row["file_name"]: row["file_md5"]
                for row in csv.DictReader(handle, delimiter="\t")
                if row.get("file_name") and row.get("file_md5")
            }

    template_lines = template.read_text(encoding="utf-8-sig").splitlines()
    if len(template_lines) < 2 or template_lines[1].split("\t") != HEADERS:
        raise SystemExit(f"Unexpected ENA template structure: {template}")

    sample_by_barcode = all_kennel_sample_map(mapping)
    asad_id_by_barcode = all_kennel_asad_id_map(
        canmag / "metadata" / "Shield Dog Samples_All Kennels_ZymoBIOMICS MagBead.xlsx"
    )
    rows: list[dict[str, str]] = []
    missing_md5: list[str] = []

    all_kennel_files = sorted(
        path for path in all_kennel_dir.glob("*.fastq.gz")
        if re.fullmatch(r"[AB]_barcode\d{2}\.fastq\.gz", path.name)
    )
    if len(all_kennel_files) != 72:
        raise SystemExit(f"Expected 72 All Kennel FASTQs, found {len(all_kennel_files)}")

    ak_protocol = (
        "Metagenomic DNA was extracted from canine fecal samples using the "
        "ZymoBIOMICS MagBead DNA/RNA workflow. Native-barcoded Oxford Nanopore "
        "ligation libraries were sequenced on PromethION."
    )
    for path in all_kennel_files:
        barcode = int(re.search(r"barcode(\d{2})", path.name).group(1))
        if barcode in EXCLUDED_ALL_KENNEL_BARCODES:
            continue
        sample_id = sample_by_barcode.get(barcode)
        if not sample_id:
            raise SystemExit(f"No sample mapping for All Kennel barcode {barcode:02d}")
        asad_id = asad_id_by_barcode.get(barcode)
        if not asad_id:
            raise SystemExit(f"No Asad sample ID for All Kennel barcode {barcode:02d}")
        checksum = cached_md5.get(path.name) or md5sum_if_available(path)
        if not checksum:
            missing_md5.append(path.name)
        rows.append(make_row(
            sample=asad_id,
            instrument="PromethION",
            library_name=f"{asad_id}_ONT_WGS",
            file_name=f"{asad_id}.fastq.gz",
            file_md5=checksum,
            protocol=ak_protocol,
        ))

    dmd_protocol = (
        "Canine fecal metagenomic DNA was extracted using the method identified "
        "in the library name (MN or Zymo HMW), prepared as a "
        "native-barcoded Oxford Nanopore ligation library, and sequenced on MinION."
    )
    for file_name, checksum in sorted(DMD_MD5.items()):
        dog = re.match(r"DMD_([^_]+)_", file_name).group(1)
        rows.append(make_row(
            sample=f"DMD_{dog}",
            instrument="MinION",
            library_name=file_name.removesuffix(".fastq.gz"),
            file_name=file_name,
            file_md5=checksum,
            protocol=dmd_protocol,
        ))

    serte_protocol = (
        "Canine fecal metagenomic DNA was extracted using the method identified "
        "in the library name (HMW or ZymoBIOMICS 96 MagBead), "
        "prepared as a native-barcoded Oxford Nanopore ligation library, and "
        "sequenced on MinION."
    )
    for file_name, checksum in sorted(SERTE_MD5.items()):
        barcode = int(re.search(r"barcode(\d{2})", file_name).group(1))
        dog, method = SERTE_BARCODE[barcode]
        rows.append(make_row(
            sample=f"Serteperti_{dog}",
            instrument="MinION",
            library_name=f"Serteperti_{dog}_{method}_ONT_WGS",
            file_name=file_name,
            file_md5=checksum,
            protocol=serte_protocol,
        ))

    if len(rows) != 84:
        raise SystemExit(f"Expected 84 manifest rows, built {len(rows)}")
    if len({row['file_name'] for row in rows}) != len(rows):
        raise SystemExit("Duplicate file_name values detected")

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        handle.write(template_lines[0] + "\n")
        writer = csv.DictWriter(handle, fieldnames=HEADERS, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows: {output}")
    print(f"Missing MD5 values: {len(missing_md5)}")
    for name in missing_md5:
        print(f"  {name}")
    if args.require_complete and missing_md5:
        raise SystemExit("Manifest is incomplete because MD5 values are missing")


if __name__ == "__main__":
    main()
